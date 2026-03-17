#COMPILER FOR EXCEL-ASM8
#USE:   py compileExcelASM8.py [program.s] ROM.xlsx

import sys
import os
import time
import math
from openpyxl import load_workbook

compiled = False;
filePath = ""
spreadsheet = ""
startTime = 0
data = []
program = []
output = []
PROGRAMLENGTH = 1024

labelOpen = False
labelToUse = ""
RED = '\033[91m';
ENDCOLOR = '\033[0m';

def integerError(lineNumber):
    print(RED + "\tInteger outside of expected range, line: " + str(lineNumber)+ ENDCOLOR)
    exit()
    
def syntaxError(lineNumber):
    print(RED + "\tSyntax Error, line: " + str(lineNumber)+ ENDCOLOR)
    exit()
    
def labelError(lineNumber):
    print(RED + "\tUnreferenced label detected before line " + str(lineNumber) + ENDCOLOR)
    exit()
    
def referenceNotFoundError(labelName):
    print(RED + "\tReference to variable or label not found, " + str(labelName) + ENDCOLOR)
    exit()

def unrecognizedError(lineNumber):
    print(RED + "\tUnrecognized Instruction, line " + str(lineNumber) + ENDCOLOR)
    exit()

def varSequenceError(lineNumber):
    print(RED + "\tVariables must be defined before program code, line " + str(lineNumber) + ENDCOLOR)
    exit()

def varUseError(lineNumber):
    print(RED + "\tVariable cannot be used like label, var: " + str(lineNumber) + ENDCOLOR)
    exit()

def orgError(lineNumber):
    print(RED + "\tProgram Count exceeds target address, line " + str(lineNumber) + ENDCOLOR)
    exit()
    
def incResourceError(resourceName, lineNumber):
    print(RED + "\tResource " + resourceName + " could not be found, line " + str(lineNumber) + ENDCOLOR)
    exit()

def lengthError(exceededWords):
    print(RED + "\tProgram length exceeds available RAM by " + str(exceededWords) + " bytes" + ENDCOLOR)
    exit()

def ROMbookError():
    print(RED + "\tCould not save to specified workbook, make sure the file is closed and try again" + ENDCOLOR)
    exit()

def ROMbookNotFoundError(spreadsheet):
    print(RED + "\tROM book " + spreadsheet + " not found" + ENDCOLOR)
    exit()
   
def labelOutOfReachError(labelName, labelLineNumber, originLine):
    print(RED + "\tLabel " + labelName + " line: " + str(labelLineNumber) + " out of branch range [-128, 127] from line: " + str(originLine) + ENDCOLOR)
    exit()
    
   
def createLine(label, operations):
    return [label, operations]

def getCurrentAddress():
    address = len(data)
    for operations in program:
        address = address + len(operations[1])
    return address

def getDistanceToLabel(labelName, originOperation):
    if (originOperation < 0):
        return None
    location = 0
    for operations in program:
        if (labelName == operations[0]):
            distance = location - originOperation
            if (distance < -128 or distance > 127):
                labelOutOfReachError(labelName, location, originOperation)
            distance = distance + 128   #reformat to signed byte
            return distance
        location = location + len(operations[1])
    return None

def getVarValue(varName):
    for var in data:
        if (varName == var[0]):
            return var[1]
    return -1    

def includeBIN(fileName):
    with open(fileName, "rb") as incFile:
        while (BYTE := incFile.read(1)):
            value = int.from_bytes(BYTE)
            program.append(createLine("", [value]))
            
    return

def parseNumber(numberString, lineNumber, isOrgInst=False):
    prefix = numberString[0]
    numberString = numberString[1:]
    result = 0
    varVal = getVarValue(prefix + numberString)      #check if references variable
    if (prefix == "$" or prefix == "@"):    #hex or address
        result = int(numberString, 16)
        if ((result > 255 and not(isOrgInst)) or (result >= PROGRAMLENGTH and isOrgInst)):
            integerError(lineNumber)
    elif (prefix == "#"):                   #decimal
        result = int(numberString)
        if ((result > 255 and not(isOrgInst)) or (result >= PROGRAMLENGTH and isOrgInst)):
            integerError(lineNumber)
    elif(not(varVal == -1)):                #is variable
        return parseNumber(varVal, lineNumber)
    else:                                   #is a label address
        result = getDistanceToLabel(prefix + numberString, -lineNumber) #negative because first pass is positive
                            #second pass is negative for calculating distance between future branches
        if (result == None and lineNumber <= -1): #second time around
            referenceNotFoundError(prefix + numberString)
        elif (result == None):
            return "LABEL-" + (prefix + numberString)            
        if ((result > 255 and not(isOrgInst)) or (result >= PROGRAMLENGTH and isOrgInst)):
            integerError(lineNumber)
    if (result < 0):
        integerError(lineNumber)
    return result

def encode(line, lineNumber):
    #convert to list of integers
    opcode = line[0]
    operand0 = 0
    twoByte = False
    #check instruction format:
    if (opcode == "BGE"):
        if (not(len(line) == 2)):
            syntaxError(lineNumber)
        twoByte = True
        operand0 = 0
        operand1 = parseNumber(line[1], lineNumber)
    elif (opcode == "LDI"):
        if (not(len(line) == 2)):
            syntaxError(lineNumber)
        twoByte = True
        operand0 = 1
        operand1 = parseNumber(line[1], lineNumber)
    elif (opcode == "CMP"):
        if (not(len(line) == 2)):
            syntaxError(lineNumber)
        twoByte = True
        operand0 = 2
        operand1 = parseNumber(line[1], lineNumber)
    elif (opcode == "ADD"):
        if (not(len(line) == 1)):
            syntaxError(lineNumber)
        operand0 = 3
    elif (opcode == "PUSH"):
        if (not(len(line) == 1)):
            syntaxError(lineNumber)
        operand0 = 4
    elif (opcode == "POP"):
        if (not(len(line) == 1)):
            syntaxError(lineNumber)
        operand0 = 5
    elif (opcode == "LDR"):
        if (not(len(line) == 1)):
            syntaxError(lineNumber)
        operand0 = 6
    elif (opcode == "STR"):
        if (not(len(line) == 1)):
            syntaxError(lineNumber)
        operand0 = 7
    elif (opcode == "CLC"):
        if (not(len(line) == 1)):
            syntaxError(lineNumber)
        operand0 = 8
    elif (opcode == "SEC"):
        if (not(len(line) == 1)):
            syntaxError(lineNumber)
        operand0 = 9
    elif (len(line) == 3 and line[1] == "="):    #variables
        if (len(program) > 0):
            varSequenceError(lineNumber)
        data.append(createLine(line[0], line[2]))
        return None
    elif (opcode == "ORG"):
        if (not(len(line) == 2)):
            syntaxError(lineNumber)
        targetAddress = parseNumber(line[1], lineNumber, True)
        currentAddress = getCurrentAddress()
        if (currentAddress > targetAddress):
            orgError(lineNumber)
        while(currentAddress < targetAddress):
            program.append(createLine("", [0]))
            currentAddress = currentAddress + 1
        return None
    elif (opcode == ".INC"):
        if (not(len(line) == 2)):
            syntaxError(lineNumber)
        line[1] = line[1].replace("\"", "")
        line[1] = line[1].replace("\'", "")
        if (not(os.path.isfile(line[1]))):
            incResourceError(line[1], lineNumber)
        includeBIN(line[1])
        return None
    else:
        unrecognizedError(lineNumber)
    
    if (not twoByte):
        return [operand0]
    else:
        return [operand0, operand1]
        
def parseProgram():
    global output
    global compiled
    for operations in program:
        for value in operations[1]:
            output.append(value)
    if (len(output) > PROGRAMLENGTH):
        lengthError(len(output) - PROGRAMLENGTH)
    compiled = True
    return

def parseUnmarkedLabels():
    pLine = 0
    instCount = 0
    for operations in program:
        valLine = 0
        for val in operations[1]:
            if ("LABEL" in str(val)):
                program[pLine][1][valLine] = parseNumber(val[6:], -instCount)
            valLine = valLine + 1
        instCount = instCount + len(operations[1])
        pLine = pLine + 1
    return

def compileASM(filepath):
    file = open(filepath, "r")
    lineNumber = 1  #file line number for specifying errors
    for line in file:
        line = line.upper()
        line = line.split(";")  #getting rid of comments
        line[0] = line[0].replace("\n", "") #removing return line
        line[0] = line[0].replace("\r", "")
        line[0] = line[0].strip()
        if (len(line[0]) > 0):
            parseLine(line[0], lineNumber)
        lineNumber = lineNumber + 1
    parseUnmarkedLabels()    
    parseProgram()
    compileResults()

def parseLine(line, lineNumber):
    global labelOpen
    global labelToUse
    labelLine = line.split(":");
    label = labelLine[0]
    if (":" in line and len(labelLine[1]) <= 1):
        if (labelOpen):
            labelError(lineNumber)
        labelToUse = label  #add a label with no operations to program
        labelOpen = True
        return
    elif (":" not in line):
        if (labelOpen):
            label = labelToUse
            labelOpen = False
        else:
            label = ""
    else:
        if (labelOpen):
            labelError(lineNumber)
        line = labelLine[1].strip()
    line = line.split(" ")
    operations = encode(line, lineNumber)
    if (not(operations == None)):
        program.append(createLine(label, operations))
    return

def sendToSpreadsheet():
    #load excel file
    workbook = load_workbook(filename = spreadsheet)     
    #open workbook
    sheet = workbook.active
    i = 0
    while (i < PROGRAMLENGTH):
        if (i < len(output)):
            sheet.cell(row = math.floor(i / 16) + 3, column = (i % 16) + 1, value = output[i])
        else:
            sheet.cell(row = math.floor(i / 16) + 3, column = (i % 16) + 1, value = 0)
        i = i + 1
        #save the file
        
    try:
        workbook.save(filename = spreadsheet)
    except:
        ROMbookError()
    return

def compileResults():
    if (not(compiled)):
        print(RED + "\tProgram could not be compiled" + ENDCOLOR)
    else:
        print("\tProgram compiled Successfully")
    print("\tProgram length in words: " + str(getCurrentAddress()))
    print("\tWriting to spreadsheet ROM...")
    sendToSpreadsheet()
    print("\tFinished in " + str(time.time()-startTime)[:6] + "s")
    exit()

if __name__ == "__main__":
    startTime = time.time()
    os.system('color')
    print("\tStarting operation...")
    
    if (len(sys.argv) == 3):
        filePath = sys.argv[1]
        spreadsheet = sys.argv[2]
    elif (len(sys.argv) == 1):
        print(RED + "\tInsufficent arguments, no ASM file specified" + ENDCOLOR)
        compileResults()
    elif (len(sys.argv) == 2):
        print(RED + "\tInsufficent arguments, no target spreadsheet specified" + ENDCOLOR)
        compileResults()
    else:
        print(RED + "\tToo Many Arguments" + ENDCOLOR)
        compileResults()
    
    if (not(os.path.isfile(spreadsheet))):
        ROMbookNotFoundError(spreadsheet)
    if (not(os.path.isfile(filePath))):
        print(RED + "\tFile " + filePath + " not found" + ENDCOLOR)
        compileResults()
    compileASM(filePath)