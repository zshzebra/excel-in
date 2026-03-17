import json
import re
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

RISC_CPU = ".reference/excelRISC-CPU/RISC-CPU.xlsx"
ROM_FILE = ".reference/excelRISC-CPU/ROM.xlsx"
OUTPUT = "cpu_model.json"


def extract_cells(wb):
    cells = {}
    functions_used = set()
    func_pattern = re.compile(r"([A-Z][A-Z0-9_.]+)\(")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                coord = f"{sheet_name}!{cell.coordinate}"
                entry = {}
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value[1:]
                    entry["formula"] = formula
                    for match in func_pattern.finditer(formula):
                        functions_used.add(match.group(1))
                    entry["cached_value"] = cell.internal_value if hasattr(cell, "internal_value") else None
                else:
                    entry["value"] = cell.value
                cells[coord] = entry

    return cells, functions_used


def extract_rom(rom_path):
    wb = load_workbook(rom_path, data_only=True)
    ws = wb.active
    rom = []
    for row_idx in range(3, 67):
        for col_idx in range(1, 17):
            val = ws.cell(row=row_idx, column=col_idx).value
            rom.append(int(val) if val is not None else 0)
    return rom


def find_circular_refs(cells):
    circular = []
    for coord, entry in cells.items():
        if "formula" not in entry:
            continue
        cell_name = coord.split("!")[-1]
        if cell_name in entry["formula"]:
            circular.append(coord)
    return circular


def main():
    print("Loading RISC-CPU.xlsx...")
    wb = load_workbook(RISC_CPU)
    cells, functions_used = extract_cells(wb)

    print("Loading ROM.xlsx...")
    rom = extract_rom(ROM_FILE)

    circular = find_circular_refs(cells)

    print(f"\nCells with formulas: {sum(1 for c in cells.values() if 'formula' in c)}")
    print(f"Cells with values: {sum(1 for c in cells.values() if 'value' in c)}")
    print(f"Functions used: {sorted(functions_used)}")
    print(f"Circular references: {len(circular)}")
    for ref in circular:
        print(f"  {ref}: {cells[ref].get('formula', '')[:80]}")

    model = {
        "cells": cells,
        "rom": rom,
        "metadata": {
            "functions_used": sorted(functions_used),
            "circular_refs": circular,
        },
    }

    with open(OUTPUT, "w") as f:
        json.dump(model, f, indent=2, default=str)

    print(f"\nExported to {OUTPUT}")


if __name__ == "__main__":
    main()
